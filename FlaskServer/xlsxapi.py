from openpyxl.workbook import Workbook
import openpyxl as xl
from utils import get_today_info, time_formatter, define_formatting_by_answer
from styles import header, default, percentage, correct, wrong, null
import os

def create_workbook():
    today = get_today_info()
    date = today['date']
    period = today['period']

    if os.path.exists(f'../Provas/{date}/{period}.xlsx'):
        return
    if not os.path.exists('../Provas'): 
        os.mkdir('../Provas')
    if not os.path.exists(f'../Provas/{date}'):
        os.mkdir(f'../Provas/{date}')

    workbook = Workbook()
    
    wbstyles = [header, default, percentage, correct, wrong, null]
    for style in wbstyles:
        workbook.add_named_style(style)

    worksheet1 = workbook.active
    worksheet1.title = 'Prova 1'
    workbook.create_sheet('Prova 2')

    first_line = ['Nome', 'Data de nascimento', 'Respostas', 'Tempo de prova', 'Quantidade de peças utilizadas', 'Tentativas', '% de acertos']
    current_column = 0
    for worksheet in workbook.worksheets:
        for i in range(len(first_line)):
            current_column += 1

            if first_line[i] == 'Respostas':
                cell = worksheet.cell(row = 1, column = current_column, value = first_line[i])
                cell.style = header
                worksheet.column_dimensions[cell.column_letter].width = len(first_line[i]) + 2
                worksheet.merge_cells(start_row = 1, start_column = current_column, end_row = 2, end_column = current_column + 4)
                current_column += 4
                continue

            cell = worksheet.cell(row = 1, column = current_column, value = first_line[i])
            cell.style = header

            worksheet.column_dimensions[cell.column_letter].width = len(first_line[i]) + 2
            if first_line[i] == 'Nome':
                worksheet.column_dimensions[cell.column_letter].width = 25

            worksheet.merge_cells(start_row = 1, start_column = current_column, end_row = 2, end_column = current_column)
        current_column = 0

    workbook.save(f'../Provas/{date}/{period}.xlsx')

def save_user_data(user_data):
    tests = [list(user_data['prova1'].values()), list(user_data['prova2'].values())]
    tests[0][2] = time_formatter(tests[0][2])
    tests[1][2] = time_formatter(tests[1][2])

    today = get_today_info()
    date = today['date']
    period = today['period']

    if not os.path.exists(f'../Provas/{date}/{period}.xlsx'):
        create_workbook()
    workbook = xl.load_workbook(f'../Provas/{date}/{period}.xlsx')

    current_index = 3
    answers_length = len(user_data['prova1']['corretas'])
    default_columns = [1, 2, 8, 9, 10, 11]
    for index, worksheet in enumerate(workbook.worksheets):
        answer_row = worksheet.max_row + 1
        worksheet.cell(row = answer_row, column = 1, value = user_data['nome']).style = 'default'
        worksheet.cell(row = answer_row, column = 2, value = user_data['nascimento']).style = 'default'
        worksheet.cell(row = answer_row, column = 8, value = tests[index][2]).style = 'default'
        worksheet.cell(row = answer_row, column = 9, value = tests[index][3]).style = 'default'
        worksheet.cell(row = answer_row, column = 10, value = tests[index][4]).style = 'default'
        worksheet.cell(row = answer_row, column = 11, value = tests[index][5]).style = 'percentage'
        for i in range(answers_length):
            worksheet.cell(
                row = answer_row,
                column = current_index,
                value = tests[index][0][i]
            ).style = 'default'
            worksheet.cell(
                row = answer_row + 1,
                column = current_index,
                value = tests[index][1][i]
            ).style = define_formatting_by_answer(tests[index][1][i], tests[index][0][i])
            current_index += 1
        for column in default_columns:
            worksheet.merge_cells(
                start_row = answer_row,
                start_column = column,
                end_row = answer_row + 1,
                end_column = column
            )
        current_index = 3
    workbook.save(f'../Provas/{date}/{period}.xlsx')
