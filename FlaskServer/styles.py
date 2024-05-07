from openpyxl.styles import NamedStyle, Border, Font, Side, PatternFill, Color, Alignment

default_side = Side(style = 'thin', color = Color(indexed = 0))

header = NamedStyle(name = 'header')
header.font = Font(bold = True, color = '00FFFFFF')
header.fill = PatternFill(start_color = Color(indexed = 30), fill_type = 'solid')
header.border = Border(left = default_side, top = default_side, right = default_side, bottom = default_side)
header.alignment = Alignment(horizontal = 'center', vertical = 'center', shrink_to_fit = False)

default = NamedStyle(name = 'default')
default.border = Border(left = default_side, top = default_side, right = default_side, bottom = default_side)
default.alignment = Alignment(horizontal = 'center', vertical = 'center', shrink_to_fit = False)

percentage = NamedStyle(name = 'percentage')
percentage.border = Border(left = default_side, top = default_side, right = default_side, bottom = default_side)
percentage.alignment = Alignment(horizontal = 'center', vertical = 'center', shrink_to_fit = False)
percentage.number_format = '0%'

correct = NamedStyle(name = 'correct')
correct.fill = PatternFill(start_color = Color(rgb = '0039E75F'), fill_type = 'solid')
correct.border = Border(left = default_side, top = default_side, right = default_side, bottom = default_side)
correct.alignment = Alignment(horizontal = 'center', vertical = 'center', shrink_to_fit = False)

wrong = NamedStyle(name = 'wrong')
wrong.fill = PatternFill(start_color = Color(rgb = '00FF7F7F'), fill_type = 'solid')
wrong.border = Border(left = default_side, top = default_side, right = default_side, bottom = default_side)
wrong.alignment = Alignment(horizontal = 'center', vertical = 'center', shrink_to_fit = False)

null = NamedStyle(name = 'null')
null.fill = PatternFill(start_color = Color(rgb = '00C0C0C0'), fill_type = 'solid')
null.border = Border(left = default_side, top = default_side, right = default_side, bottom = default_side)
null.alignment = Alignment(horizontal = 'center', vertical = 'center', shrink_to_fit = False)